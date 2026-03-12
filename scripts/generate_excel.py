#!/usr/bin/env python3
"""
Generate an audit-ready Excel workbook from SBS benchmark markdown files.

Produces one worksheet per domain with control details and columns for
auditors to record pass/fail results and notes.
"""

import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from generate_xml import (
    load_control_metadata,
    parse_markdown_file,
    strip_badge_markup,
    validate_metadata,
)

RISK_FILLS = {
    "Critical": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "High": PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid"),
    "Moderate": PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid"),
}

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

COLUMNS = [
    ("Control ID", 16),
    ("Title", 40),
    ("Control Statement", 55),
    ("Risk Level", 14),
    ("Audit Procedure", 70),
    ("Result", 12),
    ("Notes / Next Steps", 50),
]


def slugify_sheet_name(name: str) -> str:
    """Truncate and clean a category name for use as an Excel sheet name (max 31 chars)."""
    cleaned = name.strip()
    for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
        cleaned = cleaned.replace(ch, "")
    return cleaned[:31]


def build_cover_sheet(wb: Workbook, version: str, categories: dict):
    """Create an overview/cover sheet with benchmark metadata."""
    ws = wb.active
    ws.title = "Overview"

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 60

    title_font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    subtitle_font = Font(name="Calibri", size=12, color="333333")
    section_font = Font(name="Calibri", bold=True, size=12, color="1F4E79")

    row = 1
    ws.cell(row=row, column=1, value="Security Benchmark for Salesforce").font = title_font
    row += 1
    ws.cell(row=row, column=1, value="Audit Workbook").font = subtitle_font
    row += 2
    ws.cell(row=row, column=1, value="Benchmark Version:").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=2, value=version).font = Font(name="Calibri", size=11)
    row += 1
    ws.cell(row=row, column=1, value="Organization:").font = Font(name="Calibri", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value="Auditor:").font = Font(name="Calibri", bold=True, size=11)
    row += 1
    ws.cell(row=row, column=1, value="Audit Date:").font = Font(name="Calibri", bold=True, size=11)
    row += 2

    ws.cell(row=row, column=1, value="Domain Summary").font = section_font
    row += 1
    ws.cell(row=row, column=1, value="Domain").font = HEADER_FONT
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=2, value="Controls").font = HEADER_FONT
    ws.cell(row=row, column=2).fill = HEADER_FILL
    row += 1

    total = 0
    for cat_name, controls in categories.items():
        ws.cell(row=row, column=1, value=cat_name).font = Font(name="Calibri", size=11)
        ws.cell(row=row, column=2, value=len(controls)).font = Font(name="Calibri", size=11)
        total += len(controls)
        row += 1

    ws.cell(row=row, column=1, value="Total").font = Font(name="Calibri", bold=True, size=11)
    ws.cell(row=row, column=2, value=total).font = Font(name="Calibri", bold=True, size=11)
    row += 2

    ws.cell(row=row, column=1, value="Instructions").font = section_font
    row += 1
    instructions = (
        "Each tab corresponds to a benchmark domain. For every control, "
        "review the Audit Procedure, set the Result column to Pass, Fail, or N/A, "
        "and capture any findings or next steps in the Notes column."
    )
    cell = ws.cell(row=row, column=1, value=instructions)
    cell.font = Font(name="Calibri", size=10, color="555555")
    cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)


def build_domain_sheet(wb: Workbook, sheet_name: str, controls: list, metadata_dir: Path | None):
    """Create a worksheet for a single domain with one row per control."""
    ws = wb.create_sheet(title=slugify_sheet_name(sheet_name))

    pass_fail_validation = DataValidation(
        type="list",
        formula1='"Pass,Fail,N/A"',
        allow_blank=True,
    )
    pass_fail_validation.error = "Please select Pass, Fail, or N/A"
    pass_fail_validation.errorTitle = "Invalid Result"
    ws.add_data_validation(pass_fail_validation)

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes = "A2"

    for row_idx, control in enumerate(controls, start=2):
        risk_level = ""
        if metadata_dir:
            meta = load_control_metadata(control.get("id", ""), metadata_dir)
            if meta:
                validate_metadata(control.get("id", ""), meta)
                risk_level = meta.get("risk_level", "")

        values = [
            control.get("id", ""),
            control.get("title", ""),
            control.get("statement", ""),
            risk_level,
            control.get("audit_procedure", ""),
            "",  # Result — left blank for auditor
            "",  # Notes — left blank for auditor
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        risk_cell = ws.cell(row=row_idx, column=4)
        if risk_level in RISK_FILLS:
            risk_cell.fill = RISK_FILLS[risk_level]
            risk_cell.alignment = Alignment(horizontal="center", vertical="top")

        result_ref = f"F{row_idx}"
        pass_fail_validation.add(result_ref)

    ws.sheet_properties.pageSetUpPr = None


def main():
    script_dir = Path(__file__).parent
    benchmark_dir = script_dir.parent / "benchmark"
    metadata_dir = script_dir.parent / "control-metadata"
    output_file = script_dir.parent / "sbs-audit-workbook.xlsx"
    version_file = script_dir.parent / "VERSION"

    if version_file.exists():
        version = version_file.read_text().strip()
    else:
        version = "0.0.0"
        print("WARNING: VERSION file not found, using 0.0.0")

    if not metadata_dir.exists():
        metadata_dir = None

    print(f"SBS Version: {version}")
    print(f"Scanning for controls in: {benchmark_dir}")

    all_controls = []
    markdown_files = sorted(benchmark_dir.glob("*.md"))

    for md_file in markdown_files:
        print(f"Parsing {md_file.name}...")
        controls = parse_markdown_file(md_file)
        for c in controls:
            if c.get("risk"):
                c["risk"] = strip_badge_markup(c["risk"])
        all_controls.extend(controls)
        print(f"  Found {len(controls)} control(s)")

    print(f"\nTotal controls found: {len(all_controls)}")

    categories: dict[str, list] = {}
    for control in all_controls:
        cat = control.get("category", "Uncategorized")
        categories.setdefault(cat, []).append(control)

    wb = Workbook()
    build_cover_sheet(wb, version, categories)

    for cat_name, controls in categories.items():
        print(f"Building sheet: {cat_name} ({len(controls)} controls)")
        build_domain_sheet(wb, cat_name, controls, metadata_dir)

    wb.save(output_file)
    print(f"\nAudit workbook written to: {output_file}")


if __name__ == "__main__":
    main()
